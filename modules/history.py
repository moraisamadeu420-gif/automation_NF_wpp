"""
modules/history.py
Gerenciamento do histórico de notas fiscais emitidas.
Salva em CSV e exporta para Excel com formatação.
"""
import csv
from pathlib import Path
from datetime import datetime
from loguru import logger

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_DISPONIVEL = True
except ImportError:
    EXCEL_DISPONIVEL = False
    logger.warning("⚠️  pandas/openpyxl não instalados. Exportação Excel indisponível.")

from config.settings import PathConfig


# Campos do histórico
CAMPOS_CSV = [
    "data_emissao",
    "numero_nota",
    "valor",
    "valor_iss",
    "periodo",
    "tomador_cnpj",
    "tomador_nome",
    "arquivo_pdf",
    "arquivo_xml",
    "status",
    "observacao",
]


def registrar_nota(resultado: dict) -> bool:
    """
    Registra uma nota emitida no histórico CSV.
    
    Args:
        resultado: Dict retornado pelo fluxo de emissão (nfse_web.executar_emissao_nfse)
    
    Returns:
        True se registrado com sucesso
    """
    PathConfig.garantir_pastas()
    
    valor = resultado.get("valor", 0)
    aliquota = 2.0  # ISS padrão
    valor_iss = round(valor * aliquota / 100, 2)
    
    linha = {
        "data_emissao":  resultado.get("data_emissao", datetime.now().isoformat()),
        "numero_nota":   resultado.get("numero_nota", ""),
        "valor":         f"{valor:.2f}",
        "valor_iss":     f"{valor_iss:.2f}",
        "periodo":       resultado.get("periodo", ""),
        "tomador_cnpj":  "15.057.629/0001-48",  # Shopee/SPX
        "tomador_nome":  "SHOPEE COMERCIO DIGITAL DO BRASIL LTDA",
        "arquivo_pdf":   str(resultado.get("pdf", "")) or "",
        "arquivo_xml":   str(resultado.get("xml", "")) or "",
        "status":        "emitida",
        "observacao":    resultado.get("observacao", ""),
    }
    
    csv_path = PathConfig.HISTORICO_CSV
    existe = csv_path.exists()
    
    try:
        with open(csv_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=CAMPOS_CSV)
            if not existe:
                writer.writeheader()
            writer.writerow(linha)
        
        logger.success(f"✅ Nota registrada no histórico: {csv_path}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Erro ao registrar no histórico: {e}")
        return False


def registrar_falha(valor: float, periodo: str, motivo: str) -> None:
    """
    Registra uma tentativa de emissão que falhou.
    
    Args:
        valor: Valor que seria emitido
        periodo: Período de referência
        motivo: Motivo da falha
    """
    PathConfig.garantir_pastas()
    
    linha = {
        "data_emissao":  datetime.now().isoformat(),
        "numero_nota":   "",
        "valor":         f"{valor:.2f}",
        "valor_iss":     "",
        "periodo":       periodo,
        "tomador_cnpj":  "15.057.629/0001-48",
        "tomador_nome":  "SHOPEE COMERCIO DIGITAL DO BRASIL LTDA",
        "arquivo_pdf":   "",
        "arquivo_xml":   "",
        "status":        "falha",
        "observacao":    motivo[:200],
    }
    
    csv_path = PathConfig.HISTORICO_CSV
    existe = csv_path.exists()
    
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CAMPOS_CSV)
        if not existe:
            writer.writeheader()
        writer.writerow(linha)
    
    logger.warning(f"⚠️  Falha registrada no histórico: {motivo}")


def listar_historico(limite: int = 10) -> list[dict]:
    """
    Retorna as últimas N notas do histórico.
    
    Args:
        limite: Quantidade de registros a retornar
    
    Returns:
        Lista de dicts com os registros
    """
    csv_path = PathConfig.HISTORICO_CSV
    
    if not csv_path.exists():
        logger.info("📋 Histórico vazio")
        return []
    
    registros = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for linha in reader:
            registros.append(linha)
    
    return registros[-limite:]


def exportar_excel() -> Path | None:
    """
    Exporta o histórico CSV para Excel com formatação.
    
    Returns:
        Path do arquivo Excel gerado ou None se falhar
    """
    if not EXCEL_DISPONIVEL:
        logger.error("❌ pandas/openpyxl não instalados. Execute: pip install pandas openpyxl")
        return None
    
    csv_path = PathConfig.HISTORICO_CSV
    if not csv_path.exists():
        logger.warning("⚠️  Nenhum histórico encontrado para exportar")
        return None
    
    try:
        # Lê o CSV
        df = pd.read_csv(csv_path, encoding="utf-8")
        
        # Converte tipos
        df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
        df["valor_iss"] = pd.to_numeric(df["valor_iss"], errors="coerce")
        df["data_emissao"] = pd.to_datetime(df["data_emissao"], errors="coerce")
        
        # Adiciona coluna de total líquido
        df["valor_liquido"] = df["valor"] - df["valor_iss"]
        
        # Gera o Excel
        xlsx_path = PathConfig.HISTORICO_XLSX
        
        with pd.ExcelWriter(str(xlsx_path), engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Histórico NFS-e", index=False)
            
            # Formata a planilha
            wb = writer.book
            ws = writer.sheets["Histórico NFS-e"]
            
            # Cabeçalho com cor
            header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_align = Alignment(horizontal="center", vertical="center")
            
            for col_idx, col in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
            
            # Altura do cabeçalho
            ws.row_dimensions[1].height = 25
            
            # Formata linhas alternadas
            fill_claro = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
            fill_branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            borda = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )
            
            for row_idx in range(2, len(df) + 2):
                fill = fill_claro if row_idx % 2 == 0 else fill_branco
                for col_idx in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    cell.border = borda
                    cell.alignment = Alignment(vertical="center")
            
            # Auto-ajusta largura das colunas
            for col_idx, col in enumerate(df.columns, 1):
                max_len = max(
                    len(str(col)),
                    df[col].astype(str).str.len().max() if len(df) > 0 else 0
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)
            
            # Formata colunas de valor como moeda
            colunas_moeda = ["valor", "valor_iss", "valor_liquido"]
            for col_name in colunas_moeda:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    for row_idx in range(2, len(df) + 2):
                        ws.cell(row=row_idx, column=col_idx).number_format = 'R$ #,##0.00'
            
            # Linha de totais
            if len(df) > 0:
                linha_total = len(df) + 2
                ws.cell(row=linha_total, column=1, value="TOTAIS").font = Font(bold=True)
                
                for col_name in ["valor", "valor_iss", "valor_liquido"]:
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name) + 1
                        total_cell = ws.cell(row=linha_total, column=col_idx)
                        col_letter = get_column_letter(col_idx)
                        total_cell.value = f"=SUM({col_letter}2:{col_letter}{linha_total - 1})"
                        total_cell.number_format = 'R$ #,##0.00'
                        total_cell.font = Font(bold=True)
                        total_cell.fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
            
            # Segunda aba: resumo mensal
            if len(df) > 0:
                df["mes"] = df["data_emissao"].dt.to_period("M").astype(str)
                resumo = df.groupby("mes").agg(
                    total_bruto=("valor", "sum"),
                    total_iss=("valor_iss", "sum"),
                    total_liquido=("valor_liquido", "sum"),
                    qtd_notas=("numero_nota", "count"),
                ).reset_index()
                
                resumo.to_excel(writer, sheet_name="Resumo Mensal", index=False)
        
        logger.success(f"✅ Excel exportado: {xlsx_path}")
        return xlsx_path
        
    except Exception as e:
        logger.error(f"❌ Erro ao exportar Excel: {e}")
        return None


def exibir_resumo_terminal() -> None:
    """Exibe um resumo do histórico no terminal."""
    registros = listar_historico(limite=50)
    
    if not registros:
        print("📋 Nenhuma nota registrada ainda.")
        return
    
    print("\n" + "=" * 70)
    print(f"{'📊 HISTÓRICO DE NFS-e':^70}")
    print("=" * 70)
    print(f"{'Data':<20} {'Nº Nota':<12} {'Valor':>12} {'Status':<10} {'Período'}")
    print("-" * 70)
    
    total = 0.0
    for reg in registros[-10:]:  # Mostra as 10 últimas
        data = reg.get("data_emissao", "")[:10]
        nota = reg.get("numero_nota", "---")[:10]
        valor = float(reg.get("valor", 0) or 0)
        status = reg.get("status", "?")
        periodo = reg.get("periodo", "")[:20]
        total += valor if status == "emitida" else 0
        
        status_icon = "✅" if status == "emitida" else "❌"
        print(f"{data:<20} {nota:<12} R$ {valor:>9.2f}  {status_icon} {status:<8} {periodo}")
    
    print("-" * 70)
    print(f"{'TOTAL':>44} R$ {total:>9.2f}")
    print("=" * 70 + "\n")
